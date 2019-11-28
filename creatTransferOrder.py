import os

import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
from time import strftime, localtime
from tkinter import Tk, Button, Entry, Label, IntVar, Checkbutton, Radiobutton
from tkinter import filedialog

# #基础路径
# baseDir = 'C:/Users/admin/Desktop/reXuan/'
# #文件包路径
# dateName = '20191104/'
# mainName = "副本佳成仓调拨天地店11.4 test"
# writeFileName = mainName +'.xlsx'
# readFileName = mainName + '.xlsx'
# outFilepath = baseDir + dateName + '待确认出库通知单/'+strftime("%Y%m%d%H%M%S", localtime())+ writeFileName
# readFilepath = baseDir + dateName + '接受/' + readFileName

# #初始化变量，用户可改变
# outFilepath = ""
# readFilepath = ""
# #读取文件类型
# sheetType = -1
# #如果是文件类型2，门店名称其实列
# beginColForTypeTwo = -1
#################################

########常量#############
#读取文件内容字段抬头起始行
from openpyxl.styles import Font, Border, Side, Alignment

#单据创建时间
orderCreatedTime = ""
#单据主要标题
orderTitle="云商热选佳成仓库产品调拨单"
#结果单据起始行
resultBeginRow = 1
#出库类型
transType = "门店调拨"
#出库地点
outputPlace = ""
#出库联系人
outputPlaceName = "郭正峰"
#出库联系人电话
outputPlaceContact = ""
#到达时间
arriveTime = ""

#天地店
tianDiDirector = "金福王"
tianDiContact = "13777516196"
tianDiDirection = "杭州市江干区九和路28号物产天地园区一号楼一楼热选店"

#集团
jiTuanDirector = "任苏颖"
jiTuanContact = "15858112811"
jiTuanDirection = "杭州下城区环城西路56号物产中大集团一楼食堂热选店"

#国际
guoJiDirector = "屈小航"
guoJiContact = "18167116198"
guoJiDirection = "杭州市江干区凯旋路445号物产国际大厦4楼热选店"

#江干
jiangGanDirector = "毛聪聪"
jiangGanContact = "15726830081"
jiangGanDirection = "杭州市江干区庆春东路1号江干区人民政府北楼东厅热选店"

#省委党校
shengWeiDangXiaoDirector = "姚飞航"
shengWeiDangXiaoContact = "13666548755"
shengWeiDangXiaoDirection = "杭州市余杭区文一西路1000号中共浙江省委党校食堂楼旁热选店"

#上城区政府
shangChengQuDirector = "周小贞"
shangChengQuContact = "15925673320"
shangChengQuDirection = "杭州市上城区望潮路77号上城区人民政府东楼4楼食堂热选店"

#杭州市政府
hangZhouShiZhengFuDirector = "徐燕"
hangZhouShiZhengFuContact = "13958148068"
hangZhouShiZhengFuDirection = "杭州市江干区解放东路18号杭州市人民政府负1楼商业中心热选店"

#临安
linAnDirector = "项月青"
linAnContact = "138-6802-3819"
linAnDirction = "浙江省杭州市临安区衣锦街398号"

#义乌
yiWuDirector = "徐鑫"
yiWuContact = "13735318305"
yiWuDirection = "义乌市国际商贸城五区进口城1楼101号门2街60567商铺"

#页脚总计
countNum = "合计"
countBybox = "合计装箱数量（箱）"

#字段抬头
code = "国际条码"
name = "商品名称"
unit = "单位"
specification = "规格"
store = "调拨门店"
orderNum = "调拨数量"
realNum = "实际可调拨数量（仓库填写）"
supplier = "供应商"

def makeMyDir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("\\")

    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False

def file_name(file_dir):
    filesList = []
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(files)  # 当前路径下所有非目录子文件
        filesList = files
    return filesList


#获取个性化最大列数，只读取第2列，最大列数为第一个空白cell
def personal_max_col(sheet,beginRow=1):
    count = 2
    while sheet.cell(row=beginRow, column=count).value != None and\
            sheet.cell(row=beginRow, column=count).value != "None" and \
            sheet.cell(row=beginRow, column=count).value != "" :
        count = count + 1
    return count-1

#获取个性化最大行数，只读取第一列
def personal_max_row(sheet,beginCol=1):
    count = 1
    while sheet.cell(row=count, column=beginCol).value != None and\
            sheet.cell(row=count, column=beginCol).value != "None" and \
            sheet.cell(row=count, column=beginCol).value != "" :
        count = count + 1

    tmpValue = str(sheet.cell(row=count-1, column=1).value)
    if  tmpValue != None and\
            "各门店送货单请单独打印" in tmpValue:
        count = count - 1
    return count-1

#获取cell值
def getValue(sheet,row,col):
    if row == -1 or col == -1:
        return ""
    else:
        return str(sheet.cell(row=row, column=col).value)

#根据字段抬头获取位置,未找到返回-1
def getIndex(name,sheet,beginRow):
    maxCol = personal_max_col(sheet)
    index= -1
    count = 1
    while count <= maxCol:
        foundName = str(sheet.cell(row=beginRow, column=count).value)
        if name in foundName:
            index = count
            break
        count = count + 1

    return index

#门店是以title的形式出现；返回类型#Map<key=门店名，值=list<(国际条码，商品名称，单位，规格，调拨门店，调拨数量)>>
def getTypeOneInfo(sheet,beginRow):
    maxRow = personal_max_row(sheet)
    count = beginRow+1
    dictInfo = dict()
    # 获取品名col位置
    productNameIndex = getIndex("名",sheet,beginRow)
    # 获取规格col位置
    detailIndex = getIndex("规格",sheet,beginRow)
    # 获取数量col位置
    productSizeIndex = getIndex("数量",sheet,beginRow)
    # 获取条形码col位置
    codeIndex = getIndex("码",sheet,beginRow)
    # 门店位置
    shopIndex = getIndex("门店",sheet,beginRow)
    # 单位位置
    unitIndex = getIndex("单位",sheet,beginRow)

    while count <= maxRow:
        #判断门店
        tmpStoreName =  getValue(sheet,count,shopIndex)#sheet.cell(row=count, column=shopIndex).value
        if tmpStoreName not in dictInfo.keys():
            dictInfo[tmpStoreName] = list()

        dictInfo[tmpStoreName].append((
            getValue(sheet, count, codeIndex),
            getValue(sheet, count, productNameIndex),
            getValue(sheet, count, unitIndex),
            getValue(sheet, count, detailIndex),
            getValue(sheet, count, shopIndex),
            getValue(sheet, count, productSizeIndex),
        ))

        count = count + 1
    return dictInfo

#门店是以行的形式出现
def getTypeTwoInfo(sheet,beginRow,beginColToStore):
    maxRow = personal_max_row(sheet)
    maxCol = personal_max_col(sheet)
    count = beginRow+1 #数据开始行
    dictInfo = dict()
    # 获取品名col位置
    productNameIndex = getIndex("名", sheet,beginRow)
    # 获取规格col位置
    detailIndex = getIndex("规格", sheet,beginRow)
    # 获取条形码col位置
    codeIndex = getIndex("码", sheet,beginRow)
    # 单位位置
    unitIndex = getIndex("单位", sheet,beginRow)


    #获取所有门店名称，初始化list
    storeNameList = list()
    tmpCount = int(beginColToStore)
    while tmpCount <= maxCol:
        storeNameTmp = getValue(sheet,beginRow,tmpCount)
        dictInfo[storeNameTmp]=list()
        storeNameList.append(storeNameTmp)
        tmpCount = tmpCount + 1

    #获取所有数据
    while count <= maxRow:
        tmpCount = beginColToStore
        while tmpCount <= maxCol:
            productNumTmp = getValue(sheet, count, tmpCount)

            if  productNumTmp is not None and \
                    productNumTmp!= '' and\
                    productNumTmp != "None" and\
                    str(productNumTmp) != "0":

                dictInfo[storeNameList[tmpCount-beginColToStore]].append((
                    getValue(sheet, count, codeIndex),
                    getValue(sheet, count, productNameIndex),
                    getValue(sheet, count, unitIndex),
                    getValue(sheet, count, detailIndex),
                    storeNameList[tmpCount - beginColToStore],
                    productNumTmp,
                ))

            tmpCount = tmpCount + 1
        count = count + 1

    return dictInfo

#type 1: 门店是以列的形式出现； 2：门店是以行的形式出现
def getInfoByType(type,sheet,beginRow,beginColToStore=None):
    if type == 1:
        return getTypeOneInfo(sheet,int(beginRow))
    elif type == 2:
        return getTypeTwoInfo(sheet,int(beginRow),int(beginColToStore))

#根据店名获取联系人名称
def getDirectorByStore(store):
    if  "天地" in store:
        return tianDiDirector
    elif   "集团"in store:
        return jiTuanDirector
    elif   "国际"in store:
        return guoJiDirector
    elif   "江干"in store:
        return jiangGanDirector
    elif   "党校"in store:
        return shengWeiDangXiaoDirector
    elif   "上城区"in store:
        return shangChengQuDirector
    elif   "市政府"in store or "市民中心" in store:
        return hangZhouShiZhengFuDirector
    elif   "临安"in store:
        return linAnDirector
    elif "义乌" in store:
        return yiWuDirector

#根据店名获取联系方式
def getContactByStore(store):
    if   "天地"in store:
        return tianDiContact
    elif   "集团"in store:
        return jiTuanContact
    elif   "国际"in store:
        return guoJiContact
    elif   "江干"in store:
        return jiangGanContact
    elif   "党校"in store:
        return shengWeiDangXiaoContact
    elif   "上城"in store:
        return shangChengQuContact
    elif    "市政府"in store or "市民中心" in store:
        return hangZhouShiZhengFuContact
    elif   "临安"in store:
        return linAnContact
    elif "义乌" in store:
        return yiWuContact

#根据店名获取地址
def getDirectionByStore(store):
    if   "天地"in store:
        return tianDiDirection
    elif   "集团"in store:
        return jiTuanDirection
    elif   "国际"in store:
        return guoJiDirection
    elif   "江干"in store:
        return jiangGanDirection
    elif   "党校"in store:
        return shengWeiDangXiaoDirection
    elif   "上城"in store:
        return shangChengQuDirection
    elif   "市政府"in store or "市民中心"in store:
        return hangZhouShiZhengFuDirection
    elif   "临安"in store:
        return linAnDirction
    elif "义乌" in store:
        return yiWuDirection

def doWork(readFilepath,outFilepath,sheetType,beginColForTypeTwo):
    #style 准备
    # 边框style准备
    side = Side(border_style='thin', color='000000')
    border = Border(left=side,
                    right=side,
                    top=side,
                    bottom=side)

    aligmentCenter = Alignment(horizontal='center', vertical ='center')

    # 默认可读写，若有需要可以指定write_only和read_only为True
    wb = load_workbook(readFilepath)

    # 获得当前正在显示的sheet, 也可yue以用wb.get_active_sheet()
    sheet = wb.active

    #Map<key=门店名，值=list<(国际条码，商品名称，单位，规格，调拨门店，调拨数量)>>
    #type,sheet,beginRow,beginColToStore=None
    result = getInfoByType(sheetType,sheet,resultBeginRow,beginColForTypeTwo)
    print(result)

    #输出结果#####################################################################################
    newWb = Workbook()



    #创建sheet
    tmpCount = 0
    for key in result.keys():
        newSheet = newWb.create_sheet(title=key,index=tmpCount)

        #填写字段抬头
        beginRow = 1
        newSheet.cell(row=beginRow,column=1,value=code).border=border
        newSheet.cell(row=beginRow,column=2,value=name).border=border
        newSheet.cell(row=beginRow,column=3,value=unit).border=border
        newSheet.cell(row=beginRow,column=4,value=specification).border=border
        newSheet.cell(row=beginRow,column=5,value=store).border=border
        newSheet.cell(row=beginRow,column=6,value=orderNum).border=border
        newSheet.cell(row=beginRow, column=7, value=realNum).border=border

        #填写每个sheet（门店）的内容
        infoList = result[key]
        listSize = len(infoList)
        listCount = 0
        while listCount < listSize:
            newSheet.cell(row=listCount + 2, column=1, value=infoList[listCount][0]).border=border#国际条码
            newSheet.cell(row=listCount + 2, column=2, value=infoList[listCount][1]).border=border#商品名称
            newSheet.cell(row=listCount + 2, column=3, value=infoList[listCount][2]).border=border#单位
            newSheet.cell(row=listCount + 2, column=4, value=infoList[listCount][3]).border=border#规格
            newSheet.cell(row=listCount + 2, column=5, value=infoList[listCount][4]).border=border#调拨门店
            newSheet.cell(row=listCount + 2, column=6, value=infoList[listCount][5]).border=border#调拨数量
            newSheet.cell(row=listCount + 2, column=7).border = border  # 实际调拨数量
            listCount = listCount + 1


        #插入一列
        newSheet.insert_cols(1,1)
        newSheet.cell(row=1, column=1, value="序号")  # 序号
        count = 2
        while count < listCount + 2:
            newSheet.cell(row=count, column=1, value=count - 1).border=border  # 合计
            count =count + 1


        newSheet.cell(row=listCount + 2, column=1, value="合计").border=Border(left=side)  # 合计
        newSheet.cell(row=listCount + 2, column=8).border = Border(right=side)
        newSheet.cell(row=listCount + 3, column=1, value="合计装箱数（箱）").border=Border(left=side)  # 合计装箱数（箱）
        newSheet.cell(row=listCount + 3, column=8).border = Border(right=side)



        otherInfoKeyValueList=[("出库类型",transType),
                               ("出库地点",outputPlace),
                               ("出库联系人",outputPlaceName),
                               ("出库联系人电话",outputPlaceContact),
                               ("到货地点",getDirectionByStore(key)),
                               ("联系人姓名",getDirectorByStore(key)),
                                ("联系人电话",getContactByStore(key)),
                                ("要求到达时间",arriveTime)]
                               # ("调拨单号",outFilepath.split("/")[len(outFilepath.split("/")) - 1])]

        specialIndex = 4
        for item in otherInfoKeyValueList:
            newSheet.cell(row=listCount + specialIndex, column=1, value=item[0]).border=Border(left=side)
            newSheet.cell(row=listCount + specialIndex, column=8).border = Border(right=side)
            newSheet.cell(row=listCount + specialIndex, column=2, value=item[1])
            specialIndex = specialIndex + 1

        ####新增一行顶行表头
        newSheet.insert_rows(1)
        # 合并单元格,制作表头
        titleCount = 1
        while titleCount <= 8:
            newSheet.cell(row=1, column=titleCount).border = border
            titleCount  = titleCount + 1

        newSheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        newSheet.cell(row=1, column=1,value=orderTitle+orderCreatedTime).border = border
        newSheet.cell(row=1, column=1).alignment = aligmentCenter

        ######### 格式处理
        #添加底部边框
        bottomCount = 1
        maxCol = personal_max_col(newSheet,2)
        lastRowPos = listCount + specialIndex
        while bottomCount <= maxCol:
            newSheet.cell(row=lastRowPos, column=bottomCount).border = Border(bottom=side)
            bottomCount = bottomCount + 1



        #序号
        newSheet.column_dimensions['A'].width = 15
        # 国际条码
        newSheet.column_dimensions['B'].width = 20
        # 商品名称
        newSheet.column_dimensions['C'].width = 50
        # 单位
        newSheet.column_dimensions['D'].width = 15
        # 规格
        newSheet.column_dimensions['E'].width = 15
        # 调拨门店
        newSheet.column_dimensions['F'].width = 15
        # 调拨数量
        newSheet.column_dimensions['G'].width = 15
        # 实际调拨数量
        newSheet.column_dimensions['H'].width = 30

        tmpCount = tmpCount + 1


    #仓库填写
    newSheet = newWb.create_sheet(title="仓库填写",index=tmpCount)
    newSheet.cell(row=1, column=1, value="合计")
    newSheet.cell(row=1, column=2, value="箱")
    newSheet.cell(row=2, column=2, value="立方")
    newSheet.cell(row=3, column=1, value="推荐物流")
    newSheet.cell(row=4, column=1, value="推荐车型")
    newSheet.cell(row=5, column=1, value="仓库填写完毕后邮件回发至相关人员")


    newWb.remove(newWb["Sheet"])
    newWb.save(outFilepath)

def myUnion():
    outFilepath = save_path_ui.get()
    dictInfo = dict()
    finalAllInputFiles = []
    readFilepath = all_input_path_ui.get()

    if readFilepath != "":
        allFiles = file_name(readFilepath)

        for path in allFiles:
            finalAllInputFiles.append(readFilepath + path)
            dictInfo = myUnionAux(readFilepath + path,resultBeginRow,dictInfo)

    print(dictInfo)
    result = dictInfo
    #输出结果#####################################################################################
    newWb = Workbook()
    #创建sheet
    tmpCount = 0
    for key in result.keys():
        newSheet = newWb.create_sheet(title=key,index=tmpCount)

        #填写字段抬头
        beginRow = 1
        newSheet.cell(row=beginRow,column=1,value=code)
        newSheet.cell(row=beginRow,column=2,value=name)
        newSheet.cell(row=beginRow,column=3,value=unit)
        newSheet.cell(row=beginRow,column=4,value=specification)
        newSheet.cell(row=beginRow,column=5,value=store)
        newSheet.cell(row=beginRow,column=6,value=orderNum)
        newSheet.cell(row=beginRow, column=7, value=supplier)
        newSheet.cell(row=beginRow,column=8,value=realNum)

        #填写每个sheet（门店）的内容
        infoList = result[key]
        listSize = len(infoList)
        listCount = 0
        while listCount < listSize:
            newSheet.cell(row=listCount + 2, column=1, value=infoList[listCount][0])#国际条码
            newSheet.cell(row=listCount + 2, column=2, value=infoList[listCount][1])#商品名称
            newSheet.cell(row=listCount + 2, column=3, value=infoList[listCount][2])#单位
            newSheet.cell(row=listCount + 2, column=4, value=infoList[listCount][3])#规格
            newSheet.cell(row=listCount + 2, column=5, value=infoList[listCount][4])#调拨门店
            newSheet.cell(row=listCount + 2, column=6, value=infoList[listCount][5])#调拨数量
            newSheet.cell(row=listCount + 2, column=7, value=infoList[listCount][6])  # 供应商

            listCount = listCount + 1

        newSheet.cell(row=listCount + 2, column=1, value="合计")  # 合计
        newSheet.cell(row=listCount + 3, column=1, value="合计装箱数（箱）")  # 合计装箱数（箱）


        otherInfoKeyValueList=[("出库类型",transType),
                               ("出库地点",outputPlace),
                               ("出库联系人",outputPlaceName),
                               ("出库联系人电话",outputPlaceContact),
                               ("到货地点",getDirectionByStore(key)),
                               ("联系人姓名",getDirectorByStore(key)),
                                ("联系人电话",getContactByStore(key)),
                                ("要求到达时间",arriveTime),
                               ("调拨单号",outFilepath.split("/")[len(outFilepath.split("/")) - 1])]

        specialIndex = 4
        for item in otherInfoKeyValueList:
            newSheet.cell(row=listCount + specialIndex, column=1, value=item[0])
            newSheet.cell(row=listCount + specialIndex, column=2, value=item[1])
            specialIndex = specialIndex + 1


        # 格式处理
        #调整宽度
        newSheet.column_dimensions['B'].width = 20  # 国际条码
        newSheet.column_dimensions['C'].width = 50  # 商品名称
        newSheet.column_dimensions['D'].width = 15  # 单位
        newSheet.column_dimensions['E'].width = 15  # 规格
        newSheet.column_dimensions['F'].width = 15  # 调拨门店
        newSheet.column_dimensions['G'].width = 15  # 调拨数量
        newSheet.column_dimensions['H'].width = 30  # 供应商
        newSheet.column_dimensions['I'].width = 30  # 实际调拨数量

        tmpCount = tmpCount + 1


    #仓库填写
    newSheet = newWb.create_sheet(title="仓库填写",index=tmpCount)
    newSheet.cell(row=1, column=1, value="合计")
    newSheet.cell(row=1, column=2, value="箱")
    newSheet.cell(row=2, column=2, value="立方")
    newSheet.cell(row=3, column=1, value="推荐物流")
    newSheet.cell(row=4, column=1, value="推荐车型")
    newSheet.cell(row=5, column=1, value="仓库填写完毕后邮件回发至相关人员")


    newWb.remove(newWb["Sheet"])
    newWb.save(outFilepath)

def myUnionAux(readFilepath,beginRow,dictInfo):
    tmpList= readFilepath.split("/")
    orderName = tmpList[len(tmpList)-1]
    orderName = orderName.replace('.xlsx','')

    print(readFilepath)
    # 默认可读写，若有需要可以指定write_only和read_only为True
    wb = load_workbook(readFilepath)

    # 获得当前正在显示的sheet, 也可yue以用wb.get_active_sheet()
    sheet = wb.active

    # Map<key=门店名，值=list<(国际条码，商品名称，单位，规格，调拨门店，调拨数量)>>
    # type,sheet,beginRow,beginColToStore=None
    maxRow = personal_max_row(sheet)
    count = beginRow + 1

    # 获取品名col位置
    productNameIndex = getIndex("名", sheet, beginRow)
    # 获取规格col位置
    detailIndex = getIndex("规格", sheet, beginRow)
    # 获取数量col位置
    productSizeIndex = getIndex("数量", sheet, beginRow)
    # 获取条形码col位置
    codeIndex = getIndex("码", sheet, beginRow)
    # 门店位置
    shopIndex = getIndex("门店", sheet, beginRow)
    # 单位位置
    unitIndex = getIndex("单位", sheet, beginRow)

    while count <= maxRow:
        # 判断门店
        tmpStoreName = getValue(sheet, count, shopIndex)  # sheet.cell(row=count, column=shopIndex).value
        if tmpStoreName not in dictInfo.keys():
            dictInfo[tmpStoreName] = list()

        dictInfo[tmpStoreName].append((
            getValue(sheet, count, codeIndex),
            getValue(sheet, count, productNameIndex),
            getValue(sheet, count, unitIndex),
            getValue(sheet, count, detailIndex),
            getValue(sheet, count, shopIndex),
            getValue(sheet, count, productSizeIndex),
            orderName
        ))

        count = count + 1

    return dictInfo

############ui function##############
def getPath():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    print(file_path) # 打印文件的路径
    return file_path

original_table_path_ui = None

def getInputPath():
    input_path_ui.insert(0,eval(repr(getPath()).replace('\\\\', '/')))
    # 默认生成导出地址
    tmpPath = input_path_ui.get()
    tmpPathByList = tmpPath.split("/")

    #根据input的path，生成默认output
    filtPath = ""
    index = 0
    size = len(tmpPathByList)-1
    while index < size:
        filtPath = filtPath + tmpPathByList[index] + "/"
        index = index + 1


    # makeMyDir(filtPath+"/output")
    global orderCreatedTime
    orderCreatedTime = strftime("%Y%m%d%H%M%S", localtime())
    save_path_ui.insert(0,filtPath + orderCreatedTime + tmpPathByList[size])




def getAllInputPath():
    tmpPath = eval(repr(getPath()).replace('\\\\', '/'))
    tmpPathList = tmpPath.split("/")
    i = 0
    size = len(tmpPathList) - 1
    finalPath = ""
    while i < size:
        finalPath = finalPath + tmpPathList[i] + "/"
        i = i + 1

    all_input_path_ui.insert(0,finalPath)
    save_path_ui.insert(0,all_input_path_ui.get()+strftime("%Y%m%d%H%M%S", localtime())+"总表.xlsx")



def save():
    finalAllInputFiles = []
    finalAllOutPutFiles = []
    readFilepath = all_input_path_ui.get()

    if readFilepath != "":
        allFiles = file_name(readFilepath)

        for path in allFiles:
            finalAllInputFiles.append(readFilepath + path)
            finalAllOutPutFiles.append(readFilepath +  strftime("%Y%m%d%H%M%S", localtime()) + path)

        print(finalAllInputFiles)

        #循环创建所有调拨单
        size = len(finalAllInputFiles)
        count = 0
        while count < size:
            output = finalAllOutPutFiles[count]
            input = finalAllInputFiles[count]

            count = count + 1
            # 读取文件类型
            sheetType = v.get()
            # 如果是文件类型2，门店名称其实列
            beginColForTypeTwo = type_path_ui.get()

            try:
             doWork(input, output, sheetType, beginColForTypeTwo)
            except:
                print("error ===== " + input)
    else:
        # 初始化变量，用户可改变
        readFilepath = input_path_ui.get()

        outFilepath = save_path_ui.get()
        # 读取文件类型
        sheetType = v.get()
        # 如果是文件类型2，门店名称其实列
        beginColForTypeTwo = type_path_ui.get()

        doWork(readFilepath,outFilepath,sheetType,beginColForTypeTwo)

def callRB():
    print(v.get())
    if v.get()==1:
        type_path_ui['state'] = "disabled"
    else:
        type_path_ui['state'] = "normal"


root = Tk()
v=IntVar()


Label(root, text="批量创建调拨单路径: ").grid(row=0,column=0)
all_input_path_ui=Entry(root,width=60)
all_input_path_ui.grid(row=0,column=1)

baseRow = 1
baseColLeft = 0

# 通过command属性来指定Button的回调函数
Label(root, text="调拨单路径: ").grid(row=baseRow+1,column=baseColLeft)
input_path_ui=Entry(root,width=60)
input_path_ui.grid(row=baseRow+1,column=baseColLeft+1)

Label(root, text="保存路径: ").grid(row=baseRow+2,column=baseColLeft)
save_path_ui=Entry(root,width=60)
save_path_ui.grid(row=baseRow+2,column=baseColLeft+1)

Label(root, text="门店字段起始列: ").grid(row=baseRow+3,column=baseColLeft)
type_path_ui=Entry(root,width=60)
type_path_ui['state'] = "disabled"
type_path_ui.grid(row=baseRow+3,column=baseColLeft+1)

C1 = Radiobutton(root, text = "纵向类型", value="1", width = 20,command=callRB, variable=v)
C1.grid(row=baseRow+4,column=baseColLeft)
C2 = Radiobutton(root, text = "横向类型",  value="2", width = 20,command=callRB, variable=v)
C2.grid(row=baseRow+4,column=baseColLeft+1)

singlePathButton=Button(root, text='获取调拨单路径', command=getInputPath,width=15)
singlePathButton.grid(row=baseRow+5,column=baseColLeft)

allPathButton=Button(root, text='所有调拨单路径', command=getAllInputPath,width=15)
allPathButton.grid(row=baseRow+5,column=baseColLeft+1)

saveButton=Button(root, text='保存', command=save,width=15)
saveButton.grid(row=baseRow+5,column=baseColLeft+2)

saveButton=Button(root, text='合并', command=myUnion,width=15)
saveButton.grid(row=baseRow+6,column=baseColLeft)


root.mainloop()

