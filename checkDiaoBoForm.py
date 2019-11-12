#coding=gbk
import os

import pymysql
from openpyxl import load_workbook

#ֻ�Ǽ������ֶε�̧ͷ
#######################

#��ȡ���Ի����������ֻ��ȡ��һ�У��������Ϊ��һ���ֶ�Ϊ string �����е�λ��
def personal_max_row(sheet,string):
    maxRowToCheck = 5000
    count = 1
    value = sheet.cell(row=count, column=1).value
    if value is None:
        value = ""
    while count <= maxRowToCheck and string not in value:
        count = count + 1
        value = sheet.cell(row=count, column=1).value
        if value is None:
            value = ""
    return count

#��ȡ���Ի����������ֻ��ȡ��һ�У��������Ϊ��һ���հ�cell
def personal_max_col(sheet):
    count = 1
    beginRow = 1
    value = sheet.cell(row=1, column=count).value
    if value is None or "����" not in value:
        beginRow = 2

    while sheet.cell(row=beginRow, column=count).value is not None:
        count = count + 1
    return count-1

#�˷�����ѯ·���������ļ���������Ŀ¼�ļ�
def file_name(file_dir):
    toReturn = None
    for root, dirs, files in os.walk(file_dir):
        # print("=========��ǰĿ¼·��=================")
        # print(root)  # ��ǰĿ¼·��
        # print("=============ǰ·����������Ŀ¼===================")
        # print(dirs)  # ��ǰ·����������Ŀ¼
        # print("===============ǰ·�������з�Ŀ¼���ļ�=================")
        toReturn = (files) # ��ǰ·�������з�Ŀ¼���ļ�
        # toReturn = (files)
    return toReturn

#����Ƿ�item
def checkInfoFromSheetsAux(sheet,item):
    beginRow = personal_max_row(sheet, "��������")
    maxCol = personal_max_col(sheet)
    count = 1
    while count <= maxCol:
        if item in sheet.cell(row=beginRow, column=count).value:
            return True
        count = count + 1

    return  False

#����Ƿ������¼����ֶΣ� ��������",	"��Ʒ����",	"��λ",	"���",	"�����ŵ�",	"��������",	"ʵ�ʿɵ����������ֿ���д��" ���ֶ�
def checkInfoFromSheets(sheet):
    toCheckNames = ["��������",	"��Ʒ����",	"��λ",	"���",	"�����ŵ�",	"��������",	"ʵ�ʿɵ����������ֿ���д��"]

    for item in toCheckNames:
        if not checkInfoFromSheetsAux(sheet,item):
            print(item + "δ����")
            return False

    return True



root = "C:/Users/admin/Desktop/reXuan/����/�ѳɳ���"
readFilepathList = file_name(root)
index = 0
size = len(readFilepathList)

badWorkBookList = list()

#�������excel
while index < size:
    try:
        wb = load_workbook( root + "/" + readFilepathList[index])
        sheetNamesList = wb.sheetnames #��ȡ����sheet����
        if "�ֿ���д" in sheetNamesList:
            sheetNamesList.remove("�ֿ���д")
        for sheetName in sheetNamesList:
            if not checkInfoFromSheets(wb[sheetName]):
                badWorkBookList.append(root + "/" + readFilepathList[index]) #��¼Ϊ���ϱ�׼excel
        wb.close()
        index = index + 1
    except:
        print("error +++")
        badWorkBookList.append(root + "/" + readFilepathList[index])
        index = index + 1
        pass

print("============")
print(badWorkBookList)
# print(readFilepathList)
i = 0
# Ĭ�Ͽɶ�д��������Ҫ����ָ��write_only��read_onlyΪTrue

# sheet = wb.active