#coding=gbk
# import pyautogui
#
# screenWidth, screenHeight = pyautogui.size()
# print(screenWidth, screenHeight)
#
# currentMouseX, currentMouseY = pyautogui.position()
# print(currentMouseX, currentMouseY )
#
# pyautogui.moveTo(100, 150)
import os

from openpyxl import load_workbook
from openpyxl.styles import Side, Border


def file_name(file_dir):
    filesList = []
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # ��ǰĿ¼·��
        # print(dirs)  # ��ǰ·����������Ŀ¼
        # print(files)  # ��ǰ·�������з�Ŀ¼���ļ�
        filesList = files
    return filesList


#���ӱ߿�
readFilepath = "C:/Users/admin/Desktop/test.xlsx"
wb = load_workbook(readFilepath)
side = Side(border_style='thin',color='000000')
newSheet = wb.active
border = Border(left=side,
right=side,
top=side,
bottom=side)

newSheet.cell(row=1, column=1).border = border
wb.save(readFilepath)



