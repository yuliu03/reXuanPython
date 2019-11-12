#coding=gbk
import pymysql
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import uuid
from time import strftime, localtime

#��ȡ���Ի����������ֻ��ȡ��x�У��������Ϊ��һ���հ�cell
def personal_max_col(sheet,beginRow):
    count = 1
    while sheet.cell(row=beginRow, column=count).value != None:
        count = count + 1
    return count-1

#��ȡ���Ի����������ֻ��ȡ��һ�У��������Ϊ��һ���ֶ�Ϊ string
def personal_max_row(sheet,string):
    maxRowToCheck = 5000
    count = 1
    value = sheet.cell(row=count, column=1).value
    while count < maxRowToCheck and string not in str(value):
        count = count + 1
        value = sheet.cell(row=count, column=1).value
        if value == None:
            value = ""
    return count

def file_name(file_dir):
    filesList = []
    for root, dirs, files in os.walk(file_dir):
        # print(root)  # ��ǰĿ¼·��
        # print(dirs)  # ��ǰ·����������Ŀ¼
        # print(files)  # ��ǰ·�������з�Ŀ¼���ļ�
        filesList = files
    return filesList

def getProductValue(name):
    if "��������" in name:
        return "product_code"#"��������"
    elif "��Ʒ����"in name:
        return "product"#"��Ʒ����"
    elif "��λ" in name:
        return "unit"  # "��λ"
    elif "���" in name:
        return "specifications"  # "���"
    elif "�����ŵ�"in name:
        return "store_name"  # "�����ŵ�"
    elif "ʵ�ʿɵ�������"in name:
        return "real_trans_num"  # "ʵ�ʿɵ�������"
    elif "��������" in name:
        return "trans_num"  # "��������"
    return -1

def getOtherValue(name):
    if "��������" in name:
        return "output_type"#"��������"
    elif "����ص�"in name:
        return "output_place"#"����ص�"
    elif "������ϵ�˵绰" in name:
        return "output_contact"  # "������ϵ�˵绰"
    elif "������ϵ��" in name:
        return "output_person"  # "������ϵ��"
    elif "�����ص�"in name:
        return "input_place"#"�����ص�"
    elif "��ϵ������"in name:
        return "input_person"#"��ϵ������"
    elif "��ϵ�˵绰"in name:
        return "input_contact"#"��ϵ�˵绰"
    elif "Ҫ�󵽴�ʱ��"in name:
        return "arrive_time"#"Ҫ�󵽴�ʱ��"
    return -1

def getWMInfo(name):
    if "�ϼ�" in name:
        return "quantity"
    elif "����" in name:
        return "logistic"
    elif "����" in name:
        return "car_type"
    elif name is None:
        return "area_size"
    else:
        return -1

def connectDB():
    # �����ݿ�����
    db = pymysql.connect("localhost", "root", "root", "rexuan", charset='utf8')
    return db

def closeDB(db):
    # �ر����ݿ�����
    db.close()

root = "C:/Users/admin/Desktop/toInsert/"
root = "C:/Users/admin/Desktop/toInsert/test/"
filesNameList = file_name(root)


db = connectDB()
# ʹ��cursor()������ȡ�����α�
cursor = db.cursor()

for order_name in filesNameList:
    readFilepath = root + order_name
    created_time = strftime("%Y%m%d%H%M%S", localtime())

    # Ĭ�Ͽɶ�д��������Ҫ����ָ��write_only��read_onlyΪTrue
    wb = load_workbook(readFilepath)

    sheetNamesList = wb.sheetnames

    orders_id = ''.join(str(uuid.uuid4()).split('-'))
    orderSql = "INSERT INTO orders (orders_id,order_name,created_time,	quantity,logistic,area_size,car_type)VALUES	('"+orders_id+"','"+order_name+"','"+created_time+"',"


    if "�ֿ���д" in sheetNamesList:
        # print("��ʼ��ȡ�ֿ���д��Ϣ")
        actualSheet = wb["�ֿ���д"]
        maxRow = personal_max_row(actualSheet,"�ֿ���д��Ϻ��ʼ��ط��������Ա")-1
        beginRow = 1
        # print(maxRow)
        # print(beginRow)

        tmpIndex = 1
        while tmpIndex <= maxRow:
            orderSql = orderSql + "'" + str(actualSheet.cell(row=tmpIndex,column=2).value) + "',"
            tmpIndex = tmpIndex + 1

        orderSql = orderSql[:-1]+")"
        sheetNamesList.remove("�ֿ���д")

    print(orderSql)
    # print(sheetNamesList)

    # try:
    #         # print(orderSql)
    #         affectRows = cursor.execute(orderSql)
    #         # print(affectRows)
    #         db.commit()
    # except:
    #         print("==order sql error:=="+ readFilepath + "====")

    #��ȡÿ��sheet������
    for sheetName in sheetNamesList:
        stores_sheet_id = ''.join(str(uuid.uuid4()).split('-'))

        detailSql = "INSERT INTO detail_info (id,stores_sheet_id,orders_id,"
        storesSheetSql = "INSERT INTO stores_sheet(stores_sheet_id,orders_id,quantity,"
        actualSheet = wb[sheetName]

        beginRow = personal_max_row(actualSheet,"��������")
        # print(beginRow)

        maxCol = personal_max_col(actualSheet,beginRow=beginRow)

        tmpCol = 1
        while tmpCol <= maxCol:
            detailSql = detailSql + str(getProductValue(actualSheet.cell(row=beginRow, column=tmpCol).value))+","
            tmpCol = tmpCol + 1

        detailSql = detailSql[:-1]
        detailSql = detailSql + ") Values "


        #��Ʒ����Ϣ��ʵ��
        mainInfoBeginRow = beginRow + 1
        # print("mainInfoBeginRow: "+ str(mainInfoBeginRow))
        #��Ϣ�ֽ���
        separatedRow = personal_max_row(actualSheet,"��������")

        tmpPoint = separatedRow-1
        tmpInfo = str(actualSheet.cell(row=tmpPoint,column=1).value)

        #��ȡ����Ϣ���һ��λ��
        while tmpInfo is None or tmpInfo == "" or tmpInfo == "None" or "�ϼ�" in tmpInfo:
            tmpPoint = tmpPoint - 1#Ĭ����Ʒ��Ϣ���һ�к͡�����������Ϣ����һ��ֻ���һ��
            tmpInfo = str(actualSheet.cell(row=tmpPoint, column=1).value)
            if "�ϼ�װ����" in tmpInfo:
                allNumInBox = str(actualSheet.cell(row=tmpPoint, column=2).value)

        # allNumInBox = ""
        # if "�ϼ�װ����" in str(actualSheet.cell(row=tmpPoint-1,column=1).value):
        #     allNumInBox = actualSheet.cell(row=tmpPoint-1,column=2).value
        #     tmpPoint = tmpPoint - 1
        #
        # if "�ϼ�" in str(actualSheet.cell(row=tmpPoint-1,column=1).value):
        #     tmpPoint = tmpPoint - 1

        lastInfoRow = tmpPoint
        # print("last info row : " + str(lastInfoRow))

        #��ȡ��������Ϣ
        while mainInfoBeginRow <= lastInfoRow:
            tmpColPos = 1
            tmpSql = ""
            while tmpColPos <= maxCol:
                tmpSql = tmpSql + "'" + str(actualSheet.cell(row=mainInfoBeginRow,column = tmpColPos).value)+ "',"
                tmpColPos = tmpColPos + 1

            detailSql = detailSql + \
                        "('" + ''.join(str(uuid.uuid4()).split('-')) + "','"+ \
                        stores_sheet_id + "','"+ \
                        orders_id + "',"+ \
                        tmpSql[:-1] + "),"

            mainInfoBeginRow = mainInfoBeginRow + 1

        detailSql = detailSql[:-1]
        print(detailSql)

        ######��ȡ�������ݣ�����Ʒ������Ϣ��
        # print("��ȡ�������ݣ�����Ʒ������Ϣ��")
        maxRow = personal_max_row(actualSheet, "Ҫ�󵽴�ʱ��")
        otherInfoColKey = 1
        otherInfoColValue = 2
        tmpRow = separatedRow
        storesSheetValueSql = ""
        while tmpRow <= maxRow:
            storesSheetSql = storesSheetSql + str(getOtherValue(actualSheet.cell(row=tmpRow, column=otherInfoColKey).value))+","
            storesSheetValueSql = storesSheetValueSql + "'" + str(actualSheet.cell(row=tmpRow, column=otherInfoColValue).value)+"',"
            tmpRow = tmpRow + 1

        storesSheetSql = storesSheetSql[:-1]
        storesSheetSql = storesSheetSql + ") Values ('" + stores_sheet_id +"','" + orders_id + "','"+ allNumInBox + "'," + storesSheetValueSql[:-1] + ")"
        print(storesSheetSql)


        ######db example####
        #��ȡsql���ƴװ��Ϣ
        # try:
        #     affectRows = cursor.execute(storesSheetSql)
        #     # print(affectRows)
        #     print(detailSql)
        #     affectRows = cursor.execute(detailSql)
        #     # print(affectRows)
        #     result = cursor.fetchall()
        #     # print(result) #result ���ͣ� ��Ԫ���ͣ�(x1,x2,...) | xn {n = ��ѯ���ݵ������� len(x) = ÿ�����ݵ������ֶ�ֵ}
        #     db.commit()
        #
        # except:
        #     print("===="+ readFilepath + ":" + sheetName + "====")


        print("++++++++++++++++++"+ readFilepath + ":" + sheetName  +"++++++++++++++++++++")

db = closeDB(db)